import os
import random
from io import BytesIO
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ===================== 【配置】清晰缩略图 =====================
SUPPORT_FORMATS = ('.jpg', '.jpeg', '.png', '.gif', '.webp', '.avif')
IMAGE_DISPLAY_SIZE = (200, 200)  # 清晰缩略图大小（像素）
ROW_HEIGHT = int(IMAGE_DISPLAY_SIZE[1] / 1.333) + 30  # 行高适配（点，1点≈1.333像素）
COL_WIDTHS = [12, 30, int(IMAGE_DISPLAY_SIZE[0] / 7) + 3]  # 列宽适配（字符单位，1字符≈7像素）
OUTPUT_EXCEL = "女声优图鉴统计表.xlsx"

# ===================== 表格样式 =====================
HEADER_FONT = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
BODY_FONT = Font(name="微软雅黑", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="EAF1F9")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
CENTER = Alignment(horizontal="center", vertical="center")
SHEET_COLORS = ["4F81BD", "C0504D", "9BBB59", "8064A2", "F79646", "1F8A70", "3D85C6", "E06666", "6AA84F"]

# ===================== 工具函数 =====================
def get_random_image(folder_path):
    images = []
    for f in os.listdir(folder_path):
        if f.lower().endswith(SUPPORT_FORMATS):
            img_path = os.path.join(folder_path, f)
            try:
                with Image.open(img_path) as img:
                    img.verify()
                images.append(img_path)
            except:
                continue
    return random.choice(images) if images else None

def resize_image(img_path, target_size):
    """ 高质量等比例缩放，保证清晰 """
    try:
        img = Image.open(img_path).convert("RGB")
        # 🔥 高质量缩放，不模糊
        img.thumbnail(target_size, Image.Resampling.LANCZOS)
        background = Image.new("RGB", target_size, "white")
        offset = ((target_size[0] - img.width) // 2, (target_size[1] - img.height) // 2)
        background.paste(img, offset)
        return background
    except:
        return Image.new("RGB", target_size, "white")

# ===================== 主程序 =====================
if __name__ == "__main__":
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_index, item in enumerate(os.listdir("./assets")):
        project_dir = os.path.join("./assets", item)
        if not os.path.isdir(project_dir) or item.startswith("."):
            continue

        project_name = item
        print(f"正在处理企划：{project_name}")

        # 工作表名唯一化（Excel 工作表名最长 31 字符，且不能重复）
        sheet_title = project_name[:30]
        suffix = 2
        while sheet_title in wb.sheetnames:
            sheet_title = project_name[:29] + f"-{suffix}"
            suffix += 1

        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_properties.tabColor = SHEET_COLORS[sheet_index % len(SHEET_COLORS)]

        # ---------- 表头 ----------
        ws.append(["编号", "女声优名字", "代表图片"])
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.alignment = CENTER
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 30

        # 收集声优文件夹
        seiyu_dirs = []
        for subfolder in os.listdir(project_dir):
            sf_path = os.path.join(project_dir, subfolder)
            if os.path.isdir(sf_path) and '-' in subfolder:
                seiyu_dirs.append((subfolder, sf_path))

        # 按编号排序
        try:
            seiyu_dirs.sort(key=lambda x: int(x[0].split('-')[0]))
        except:
            pass

        # ---------- 写入数据 ----------
        row = 2
        for folder_name, folder_path in seiyu_dirs:
            parts = folder_name.split('-', 1)
            sid = parts[0].strip()
            name = parts[1].strip() if len(parts) > 1 else "未知"

            cell_no = ws.cell(row=row, column=1, value=sid)
            cell_name = ws.cell(row=row, column=2, value=name)
            for cell in (cell_no, cell_name):
                cell.alignment = CENTER
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if row % 2 == 0:
                    cell.fill = ZEBRA_FILL

            # 插入清晰缩略图（BytesIO 内存处理，不产生临时文件）
            img_path = get_random_image(folder_path)
            if img_path:
                resized_img = resize_image(img_path, IMAGE_DISPLAY_SIZE)
                buf = BytesIO()
                resized_img.save(buf, format="JPEG", quality=95)
                buf.seek(0)
                xl_img = XLImage(buf)
                ws.add_image(xl_img, f"C{row}")

            ws.row_dimensions[row].height = ROW_HEIGHT
            row += 1

        # ---------- 表格美化 ----------
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"  # 固定表头，长列表滚动时始终可见
        if row > 2:
            ws.auto_filter.ref = f"A1:C{row - 1}"  # 表头启用筛选

    wb.save(OUTPUT_EXCEL)
    print(f"\n导出完成：{OUTPUT_EXCEL}")
